import json
path = r"D:/SEM 6/ML Project/Implementation/Datasets/apy.xlsx"

import pandas as pd
import numpy as np
import xlrd
import openpyxl

wbk = openpyxl.load_workbook(path) 
sheet = wbk.active


wb = xlrd.open_workbook(r"D:\SEM 6\Data Visualization Project\Datasets\Lat_long_data.xlsx") 
sheet_lat = wb.sheet_by_index(0)

loc_data = {};
for i in range(1,sheet_lat.nrows):
    dis = sheet_lat.cell(i,1).value
    temp = sheet_lat.cell(i,0).value + "_" + dis.upper()
    lat = sheet_lat.cell(i,2).value;
    long = sheet_lat.cell(i,3).value;
    loc_data[temp] = [lat,long]
print(loc_data)

for i in range(2,sheet.max_row+1): 
    state = sheet.cell(row=i, column=1).value
    dis = sheet.cell(row = i, column = 2).value
    key = state + "_" + dis
    sheet.cell(row = i, column = 5).value = loc_data[key][0]
    sheet.cell(row = i, column = 6).value = loc_data[key][1]
    print(i)
    
wbk.save(path)
wbk.close    











'''
rain_data = pd.read_csv(r"D:/SEM 6/ML Project/Implementation/Datasets/csv data/rainfall in india 1901-2015.csv")
rain_data = np.array(rain_data)
rain_info = {}

for k in rain_data:
    if(k[1]>=1997):
        for state in k[0].split('+'):
            r = state + "_" + str(int(k[1])) + "_" + "Rabi"
            rain_info[r] = []
            r = state + "_" + str(int(k[1])) + "_" + "Autumn"
            rain_info[r] = []
            r = state + "_" + str(int(k[1])) + "_" + "Kharif"
            rain_info[r] = []
            r = state + "_" + str(int(k[1])) + "_" + "Summer"
            rain_info[r] = []
            r = state + "_" + str(int(k[1])) + "_" + "Winter"
            rain_info[r] = []
            r = state + "_" + str(int(k[1])) + "_" + "Whole Year"
            rain_info[r] = []

for k in rain_data:
    if(k[1]>=1997):
        for state in k[0].split('+'):        
            r = state + "_" + str(int(k[1])) + "_" + "Rabi"
            rabi = (k[2] + k[3] + k[4] + k[11] + k[12] + k[13])/6
            rain_info[r].append(rabi)
            
            autumn = (k[10] + k[11] + k[12] + k[13])/4
            a = state + "_" + str(int(k[1])) + "_" + "Autumn"
            rain_info[a].append(autumn)
            
            kharif = (k[8] + k[9] + k[10] + k[11])/4
            a = state + "_" + str(int(k[1])) + "_" + "Kharif"
            rain_info[a].append(kharif)
            
            summer = (k[6] + k[7] + k[8] + k[9] + k[10])/5
            a = state + "_" + str(int(k[1])) + "_" + "Summer"
            rain_info[a].append(summer)
            
            winter = (k[2] + k[3] + k[4] + k[13])/4
            a = state + "_" +str(int(k[1])) + "_" + "Winter"
            rain_info[a].append(winter)
            
            a = state + "_" + str(int(k[1])) + "_" + "Whole Year"
            rain_info[a].append(k[14]/12)

for k in rain_info:
    rain_info[k] = sum(rain_info[k])/len(rain_info[k])
print(rain_info)


for i in range(2,sheet.max_row+1): 
    state = sheet.cell(row=i, column=1).value
    year = sheet.cell(row = i, column = 3).value
    season = sheet.cell(row = i, column = 4).value 
    key = state + "_" + str(int(year)) + "_" + season
    sheet.cell(row = i, column = 5).value = rain_info[key]
    print(i)
    
wbk.save(path)
wbk.close 

'''
















'''
temp_data = pd.read_csv(r"D:/SEM 6/ML Project/Implementation/Datasets/csv data/temperature data.csv")
temp_data = np.array(temp_data)
temp_info = {}

for k in temp_data:
    if(k[0]>=1997):
        
        r = str(int(k[0])) + "_" + "Rabi"
        rabi = (k[1] + k[2] + k[3] + k[10] + k[11] + k[12])/6
        temp_info[r] = rabi
        
        autumn = (k[9] + k[10] + k[11] + k[12])/4
        a = str(int(k[0])) + "_" + "Autumn"
        temp_info[a] = autumn
        
        kharif = (k[7] + k[8] + k[9] + k[10])/4
        a = str(int(k[0])) + "_" + "Kharif"
        temp_info[a] = kharif
        
        summer = (k[5] + k[6] + k[7] + k[8] + k[9])/5
        a = str(int(k[0])) + "_" + "Summer"
        temp_info[a] = summer
        
        winter = (k[1] + k[2] + k[3] + k[12])/4
        a = str(int(k[0])) + "_" + "Winter"
        temp_info[a] = winter
        
        a = str(int(k[0])) + "_" + "Whole Year"
        temp_info[a] = k[13]
print(temp_info)
        
        
for i in range(2,sheet.max_row+1): 
    year = sheet.cell(row = i, column = 3).value
    season = sheet.cell(row = i, column = 4).value 
    key = str(int(year)) + "_" + season
    sheet.cell(row = i, column = 5).value = temp_info[key]
    print(i)
    
wbk.save(path)
wbk.close   
'''
 
'''

wks.cell(row=i+1, column=3).value = lat
wks.cell(row=i+1, column=4).value = long

'''

'''
state_list = {}

for k in data:
    if k[0].strip() not in state_list:
        state_list[k[0].strip()] = {}

for state in state_list:
    for k in data:
        if((state == k[0].strip()) and (k[1].strip() not in state_list[state])):
            state_list[state][k[1].strip()] = {}

count = 0
for state in state_list:
    for district in state_list[state]:
        for k in data:
            if((state == k[0].strip()) and (district == k[1].strip()) and not(np.isnan(k[5]) or np.isnan(k[6]))):
                if (k[4].strip() not in state_list[state][district]):
                    state_list[state][district][k[4].strip()] = [[k[2],k[3].strip(),k[5],k[6]]]
                    count+=1
                else:
                    state_list[state][district][k[4].strip()].append([k[2],k[3].strip(),k[5],k[6]])
                    count+=1
# print(state_list)
print(count)
    
with open(r"./assets/datasets/crops.json",'w') as json_file:
    json.dump(state_list,json_file)
'''