path = r"D:\SEM 6\Data Visualization Project\Datasets\apy.csv"

import pandas as pd
import numpy as np 
import requests
import xlrd
import openpyxl
import json

wb = xlrd.open_workbook(r"D:\SEM 6\ML Project\Implementation\Datasets\apy.xlsx") 

sheet = wb.sheet_by_index(0)


data = {}
for i in range(1,sheet.nrows):
    if(sheet.cell(i,9)=='' or sheet.cell(i,9)==0 or sheet.cell(i,10)=='' or sheet.cell(i,10)==0):
        pass
    else:
        season = sheet.cell(i,3).value
        crop = sheet.cell(i,8).value
        if '/' in crop:
            crop = crop.replace('/','&')
        if(season not in data):
            data[season] = []
        if crop not in data[season]:
            data[season].append(crop)
print(data)
json_object = json.dumps(data, indent = 4) 
with open("SeasonwiseCrops.json", "w") as outfile: 
    outfile.write(json_object)





'''
data = {};
for i in range(1,sheet.nrows):
    
    state = sheet.cell(i,0).value;
    dis = sheet.cell(i,1).value;
    lat = sheet.cell(i,2).value;
    long = sheet.cell(i,3).value;
    if(state not in data):
        data[state] = {}
    if(dis not in data[state]):
        data[state][dis] = []
    
    
    data[state][dis] = [lat,long]
    
json_object = json.dumps(data, indent = 4) 
with open("state_dis_loc.json", "w") as outfile: 
    outfile.write(json_object)
'''











'''
wb = xlrd.open_workbook(r"D:\SEM 6\Data Visualization Project\Datasets\Lat_long_data.xlsx") 
sheet = wb.sheet_by_index(0)

data = [];
for i in range(1,sheet.nrows):
    temp = {}
    temp["state"] = sheet.cell(i,0).value;
    temp["dis"] = sheet.cell(i,1).value;
    temp["lat"] = sheet.cell(i,2).value;
    temp["long"] = sheet.cell(i,3).value;
    data.append(temp)
    
with open(r"D:\SEM 6\Data Visualization Project\Datasets\lat_long.json",'w') as json_file:
    json.dump(data,json_file)








wbkName = r"D:\SEM 6\Data Visualization Project\Datasets\Lat_long_data.xlsx"


for i in range(600,sheet.nrows):
    url = "https://api.opencagedata.com/geocode/v1/json?q="
    url = url + sheet.cell_value(i,1) + "+" + sheet.cell_value(i,0) + "+" + "India" + "&key=5f831da67b1e449096e98f70e65ee22b"
    response = requests.get(url)
    print(response.status_code)
    if(response.status_code == 200):
        lat = response.json()['results'][0]['geometry']['lat']
        long = response.json()['results'][0]['geometry']['lng']
   
        wbk = openpyxl.load_workbook(wbkName)
        wks = wbk.active
        wks.cell(row=i+1, column=3).value = lat
        wks.cell(row=i+1, column=4).value = long
        wbk.save(wbkName)
    
        wbk.close

'''












'''
data = pd.read_csv(path)
data = np.array(data)
state_dis_list = []

for k in data:
    temp = [k[0].strip(),k[1].strip().title()]
    if temp not in state_dis_list:
        state_dis_list.append(temp)

df = pd.DataFrame(state_dis_list)
df.to_excel(r"D:\SEM 6\Data Visualization Project\Datasets\Lat_long_data.xlsx",index=False)
'''









'''
for k in state_dis_list:
    url = "https://api.opencagedata.com/geocode/v1/json?q="
    url = url + k[1] + "+" + k[0] + "+" + "India" + "&key=5f831da67b1e449096e98f70e65ee22b"
    response = requests.get(url)
    lat = response.json()['results'][0]['geometry']['lat']
    k.append(lat)
    long = response.json()['results'][0]['geometry']['lng']
    k.append(long)
    

response = requests.get("https://api.opencagedata.com/geocode/v1/json?q=Vadodara+Gujarat+India&key=5f831da67b1e449096e98f70e65ee22b")
lat = response.json()
long = response.json()['results'][0]['geometry']['lng']
print(lat,long)
'''

